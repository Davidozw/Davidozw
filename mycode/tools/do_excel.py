"""
File:    do_excel
Author:  ouzongwei
Time:   2021/3/14 23:47
Ide_name:PyCharm
"""
from openpyxl import load_workbook
from common.path_file import *
from tools.do_conf import Do_conf
from common.G_data import Data_common

class Do_excel:


    #用静态方法获取excel表中的数据
    @staticmethod
    def getdata(filename):
        wb=load_workbook(filename)
        item = eval(Do_conf().config_data(conf_path,getattr(Data_common,'section'),getattr(Data_common,'option')))
        test_data=[]
        for key in item:
            sheetname=wb[key]
            if item[key] == 'all':
                for i in range(2,sheetname.max_row+1):
                    raw_data={}
                    raw_data['case_id'] = sheetname.cell(i,1).value
                    raw_data['url'] = sheetname.cell(i, 2).value
                    raw_data['data'] = sheetname.cell(i, 3).value
                    raw_data['method'] = sheetname.cell(i, 4).value
                    raw_data['expected'] = sheetname.cell(i, 5).value
                    raw_data['sheet_name'] = key
                    test_data.append(raw_data)
            else:
                for j in item[key]:
                    raw_data = {}
                    raw_data['case_id'] = sheetname.cell(j+1, 1).value
                    raw_data['url'] = sheetname.cell(j+1, 2).value
                    raw_data['data'] = sheetname.cell(j+1, 3).value
                    raw_data['method'] = sheetname.cell(j+1, 4).value
                    raw_data['expected'] = sheetname.cell(j+1, 5).value
                    raw_data['sheet_name'] = key
                    test_data.append(raw_data)
        return test_data

    #用静态方法将数据写入文件
    @staticmethod
    def write_back(filename,sheet,row,column,message):
        wb = load_workbook(filename)
        sheetname = wb[sheet]
        sheetname.cell(row,column).value = message
        wb.save(filename)





if __name__ == '__main__':
    dd=Do_excel.getdata(data_path)
    Do_excel.write_back(data_path,'python',6,1,111)